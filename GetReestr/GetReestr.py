""" Утилита по загрузке данных по госреестровым ценам и сохранении в файл для ПК Аптека-Урал """
import datetime
import hashlib
import os
import shutil
import zipfile

import urllib3
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet

# import requests
import Defs
import Mail


def from_row_in_tuple(from_row):
    res = []
    for val in from_row:
        res = res + [val.value]
    while len(res) < 11:
        res = res + ['']
    return res


def file_as_bytes(file):
    with file:
        return file.read()


def load_file(start_page, zip_file):
    """Загрузка реестра с сайта"""
    http = urllib3.PoolManager()
    response = http.request("GET", start_page)
    html_start_page = str(response.data)

    """Поиск файла на сайте"""
    link_start = html_start_page.find('GetLimPrice.ashx?')
    link_end = html_start_page.find('\\', link_start)
    link = html_start_page[link_start:link_end]

    """Удаляем старый файл"""
    if os.path.exists(zip_file):
        os.remove(zip_file)
    zip_url = 'https://grls.rosminzdrav.ru/' + link
    response = http.request("GET", zip_url)

    print(str(len(f"Размер файла: {response.data}")))

    f = open(zip_file, 'wb')
    f.write(response.data)
    f.close()


def unpack_zip(zip_file, tmp_path, data_file):
    """Распаковка файла"""
    stories_zip = zipfile.ZipFile(zip_file)
    for file in stories_zip.namelist():
        stories_zip.extract(file, tmp_path)
        os.rename(tmp_path + file, data_file)
    stories_zip.close()


def create_file(result_tmp, columns):
    """Создаем и формируем структуру файла"""
    f = open(result_tmp, 'w', encoding="cp1251")
    line = ""
    for key, value in columns.items():
        if int(value) > -2:
            if key != 'DATA_REG_UDOST':
                line = line + str(key) + "\t"
    line = line + str("NAME_SHEET") + "\t"
    if line[-1] == "\t":
        line = line[:-1]
    f.write(line + "\n")
    return f


def parse_sheet(rb: Workbook, num_sheet, start_line, columns, formats, file, name_sheet):
    sheet: ReadOnlyWorksheet = rb.worksheets[num_sheet]
    """Ищем последнюю запись на странице"""
    max_idx = sheet.max_row
    """Заполняем данные"""
    idx = start_line
    errors = ""
    last_date = ""

    for row in sheet.iter_rows(min_row=start_line, max_row=max_idx):
        values = from_row_in_tuple(row)

        if idx % 500 == 0:
            print(f'Загружено: {idx} / {max_idx - start_line}')

        line = ""
        for key, value in columns.items():

            if int(value) > -2:
                if int(value) == -1:
                    if key == 'DATE_FROM':
                        line = line + "01.01.2000\t"
                    else:
                        line = line + "\t"
                else:
                    if formats[key] == 'D':
                        # читаем дату
                        time = values[int(value)]
                        s = time.strftime("%d.%m.%Y")

                        if key == 'DATE_FROM':
                            if idx == start_line:
                                last_date = time
                            else:
                                if last_date < time:
                                    last_date = time
                    else:
                        if (key != 'DATA_REG_UDOST'):
                            s = str(values[int(value)])
                        else:
                            s = values[9].strip()
                            s = s[0:10]

                    s = s.replace("\n", " ").replace("\r", " ").replace("   ", " ").replace("None","")
                    """Временно для теста, пока не добавим поля в АУ"""
                    if (key == 'NUM_REG_UDOST'):
                        line = line + name_sheet + ' ' + s + " от "
                    else:
                        line = line + s + "\t"

        try:
            line = line + name_sheet + "\t"
            if line[-1] == "\t":
                line = line[:-1]
            file.write(line + "\n")
        except Exception as e:
            errors = errors + str(idx) + ": \n" + line + " \n   (" + str(e) + ")\n"
            print(line)

        idx = idx + 1

    print(idx)
    print("End loading")
    return file, errors, last_date


def parse_xls(data_file, file_parse_params, file):
    """Открываем файл"""
    wb = load_workbook(data_file, read_only=True, data_only=True)
    """Парсим страницы"""
    all_errors = ""
    last_date = ""
    for sheet in file_parse_params:
        file, errors, last_date = parse_sheet(
            wb,
            int(sheet["sheet"]["num_sheet"]),
            int(sheet["sheet"]["start_line"]),
            sheet["columns"], sheet["formats"],
            file, str(sheet["sheet"]["name_sheet"]))
        all_errors = all_errors + errors
    wb.close()
    return all_errors, last_date


def write_md5_to_file(file_path: str, md5_value: str) -> None:
    with open(file_path, 'w+') as file:
        file.write(md5_value)


def get_md5hash(result_file, result_tmp, md5_file):
    old_hash = ''
    if os.path.exists(result_file):
        old_hash = hashlib.md5(open(result_file, 'rb').read()).hexdigest()
    new_hash = hashlib.md5(open(result_tmp, 'rb').read()).hexdigest()
    write_md5_to_file(md5_file, new_hash)

    new_file = 0
    if old_hash != '':
        if old_hash != new_hash:
            new_file = 1
    else:
        new_file = 1
    return new_file


def replace_file(new_file, tmp_path, result_file, copy_file):
    zip_to_send = tmp_path + "reestr.zip"
    if new_file == 1:
        z = zipfile.ZipFile(zip_to_send, 'w', zipfile.ZIP_DEFLATED)
        z.write(result_file, 'reestr.txt')
        z.close()
        # копирование файлов
        for file in copy_file:
            shutil.copy(zip_to_send, file)
    return zip_to_send


def send_email(send_by_email, new_file, emails, last_date, errors, zip_to_send):
    if (send_by_email == '1') and (new_file == 1):
        msg_subj = emails['subject']
        last_date_str = last_date.strftime("%Y-%m-%d")
        msg_subj = msg_subj.replace("[LastData]", last_date_str)

        msg_text = "Новый реестр загружен \n"
        if errors != "":
            msg_text = msg_text + "Errors: \n" + errors
        Mail.send_email(emails['host'], emails['email_addr'], emails['password'],
                        emails['addr_to'], msg_subj, msg_text, [zip_to_send])


def main():
    """Загрузка параметров"""
    params = Defs.Params()

    """Загрузка файла"""
    load_file(params.start_page, params.zip_file)

    """Извлечение файла"""
    unpack_zip(params.zip_file, params.tmp_path, params.xlsx_file)

    """Создаем и формируем структуру файла"""
    file = create_file(params.result_tmp, params.file_parse_params[0]["columns"])

    """Открытие файла XLS и парсим"""
    errors, last_date = parse_xls(params.xlsx_file, params.file_parse_params, file)

    """Закрываем файл"""
    file.close()

    """Удаляем временные файлы"""
    os.remove(params.xlsx_file)
    os.remove(params.zip_file)

    """Считаем md5 старого и нового файла"""
    new_file = get_md5hash(params.result_file, params.result_tmp, params.md5_file)

    if os.path.exists(params.result_file):
        os.remove(params.result_file)
    os.rename(params.result_tmp, params.result_file)

    """Замена файлов, создание архива и копирование"""
    zip_to_send = replace_file(new_file, params.tmp_path, params.result_file, params.copy_file)

    """Отправляем данные по почте"""
    send_email(params.send_by_email, new_file, params.emails, last_date, errors, zip_to_send)


if __name__ == '__main__':
    main()
